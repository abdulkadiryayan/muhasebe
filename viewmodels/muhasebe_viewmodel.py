from models.database import MuhasebeDB
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

class MuhasebeViewModel:
    def __init__(self):
        self.db = MuhasebeDB()
        
    # Cari Hesap İşlemleri
    def cari_hesap_ekle(self, musteri_adi: str, borc: float = 0, alacak: float = 0, islem_tarihi = None) -> bool:
        try:
            self.db.cari_hesap_ekle(musteri_adi, borc, alacak, islem_tarihi)
            return True
        except Exception as e:
            print(f"Hata: Cari hesap eklenirken bir sorun oluştu - {str(e)}")
            return False
            
    def cari_hesap_listele(self):
        try:
            self.db.cursor.execute("SELECT * FROM cari_hesap")
            return self.db.cursor.fetchall()
        except Exception as e:
            print(f"Hata: Cari hesaplar listelenirken bir sorun oluştu - {str(e)}")
            return []
            
    def cari_hesap_bakiye_hesapla(self, id: int) -> float:
        try:
            self.db.cursor.execute(
                "SELECT (alacak - borc) as bakiye FROM cari_hesap WHERE id=?", 
                (id,)
            )
            sonuc = self.db.cursor.fetchone()
            return sonuc[0] if sonuc else 0
        except Exception as e:
            print(f"Hata: Bakiye hesaplanırken bir sorun oluştu - {str(e)}")
            return 0

    # Cari Hesap İşlemleri - Silme ve Güncelleme
    def cari_hesap_sil(self, id: int) -> bool:
        try:
            self.db.cari_hesap_sil(id)
            return True
        except Exception as e:
            print(f"Hata: Cari hesap silinirken bir sorun oluştu - {str(e)}")
            return False
            
    def cari_hesap_guncelle(self, id: int, musteri_adi: str = None, 
                           borc: float = None, alacak: float = None) -> bool:
        try:
            self.db.cari_hesap_guncelle(id, musteri_adi, borc, alacak)
            return True
        except Exception as e:
            print(f"Hata: Cari hesap güncellenirken bir sorun oluştu - {str(e)}")
            return False

    # Kasa İşlemleri
    def kasa_ekle(self, kasa_adi: str, gelir: float = 0, gider: float = 0, islem_tarihi = None) -> bool:
        try:
            self.db.kasa_ekle(kasa_adi, gelir, gider, islem_tarihi)
            return True
        except Exception as e:
            print(f"Hata: Kasa eklenirken bir sorun oluştu - {str(e)}")
            return False
            
    def kasa_bakiye_hesapla(self, id: int) -> float:
        try:
            self.db.cursor.execute(
                "SELECT (gelir - gider) as bakiye FROM kasa_yonetimi WHERE id=?", 
                (id,)
            )
            sonuc = self.db.cursor.fetchone()
            return sonuc[0] if sonuc else 0
        except Exception as e:
            print(f"Hata: Kasa bakiyesi hesaplanırken bir sorun oluştu - {str(e)}")
            return 0

    # Kasa İşlemleri - Silme ve Güncelleme
    def kasa_sil(self, id: int) -> bool:
        try:
            self.db.kasa_sil(id)
            return True
        except Exception as e:
            print(f"Hata: Kasa silinirken bir sorun oluştu - {str(e)}")
            return False
            
    def kasa_guncelle(self, id: int, kasa_adi: str = None, 
                      gelir: float = None, gider: float = None) -> bool:
        try:
            self.db.kasa_guncelle(id, kasa_adi, gelir, gider)
            return True
        except Exception as e:
            print(f"Hata: Kasa güncellenirken bir sorun oluştu - {str(e)}")
            return False

    def kasa_listele(self):
        try:
            self.db.cursor.execute("SELECT * FROM kasa_yonetimi")
            return self.db.cursor.fetchall()
        except Exception as e:
            print(f"Hata: Kasa listesi alınırken bir sorun oluştu - {str(e)}")
            return []

    # Fatura İşlemleri
    def fatura_ekle(self, fatura_no: str, tutar: float, tur: str, islem_tarihi = None) -> bool:
        try:
            self.db.fatura_ekle(fatura_no, tutar, tur, islem_tarihi)
            return True
        except Exception as e:
            print(f"Hata: Fatura eklenirken bir sorun oluştu - {str(e)}")
            return False
            
    def fatura_listele(self, tur: str = None):
        try:
            if tur:
                self.db.cursor.execute("SELECT * FROM faturalar_irsaliyeler WHERE tur=?", (tur,))
            else:
                self.db.cursor.execute("SELECT * FROM faturalar_irsaliyeler")
            return self.db.cursor.fetchall()
        except Exception as e:
            print(f"Hata: Faturalar listelenirken bir sorun oluştu - {str(e)}")
            return []

    # Fatura İşlemleri - Silme ve Güncelleme
    def fatura_sil(self, id: int) -> bool:
        try:
            self.db.fatura_sil(id)
            return True
        except Exception as e:
            print(f"Hata: Fatura silinirken bir sorun oluştu - {str(e)}")
            return False
            
    def fatura_guncelle(self, id: int, fatura_no: str = None, 
                       tutar: float = None, tur: str = None) -> bool:
        try:
            self.db.fatura_guncelle(id, fatura_no, tutar, tur)
            return True
        except Exception as e:
            print(f"Hata: Fatura güncellenirken bir sorun oluştu - {str(e)}")
            return False

    # Çek/Senet İşlemleri
    def cek_senet_ekle(self, evrak_turu: str, vade_tarihi: str, tutar: float) -> bool:
        try:
            self.db.cek_senet_ekle(evrak_turu, vade_tarihi, tutar)
            return True
        except Exception as e:
            print(f"Hata: Çek/Senet eklenirken bir sorun oluştu - {str(e)}")
            return False
            
    def vadesi_yaklasan_cek_senetler(self, gun_sayisi: int = None):
        try:
            if gun_sayisi:
                # Vadesi yaklaşanları listele
                bugun = datetime.now().strftime('%Y-%m-%d')
                self.db.cursor.execute("""
                    SELECT * FROM cek_senet 
                    WHERE date(vade_tarihi) BETWEEN date(?) AND date(?, '+' || ? || ' days')
                    ORDER BY vade_tarihi
                """, (bugun, bugun, gun_sayisi))
            else:
                # Tüm çek/senetleri listele
                self.db.cursor.execute("SELECT * FROM cek_senet ORDER BY vade_tarihi")
                
            return self.db.cursor.fetchall()
        except Exception as e:
            print(f"Hata: Çek/senetler listelenirken bir sorun oluştu - {str(e)}")
            return []

    # Çek/Senet İşlemleri - Silme ve Güncelleme
    def cek_senet_sil(self, id: int) -> bool:
        try:
            self.db.cek_senet_sil(id)
            return True
        except Exception as e:
            print(f"Hata: Çek/Senet silinirken bir sorun oluştu - {str(e)}")
            return False
            
    def cek_senet_guncelle(self, id: int, evrak_turu: str = None, 
                          vade_tarihi: str = None, tutar: float = None) -> bool:
        try:
            self.db.cek_senet_guncelle(id, evrak_turu, vade_tarihi, tutar)
            return True
        except Exception as e:
            print(f"Hata: Çek/Senet güncellenirken bir sorun oluştu - {str(e)}")
            return False 

    def excel_export(self, dosya_yolu: str = None) -> tuple[bool, str]:
        """Tüm verileri Excel dosyasına aktarır"""
        try:
            wb = Workbook()
            
            # Varsayılan sheet'i sil
            wb.remove(wb.active)
            
            # Cari Hesaplar sayfası
            ws_cari = wb.create_sheet("Cari Hesaplar")
            ws_cari.append(["ID", "Müşteri Adı", "Borç", "Alacak", "Bakiye"])
            
            for hesap in self.cari_hesap_listele():
                bakiye = self.cari_hesap_bakiye_hesapla(hesap[0])
                ws_cari.append([hesap[0], hesap[1], hesap[2], hesap[3], bakiye])
            
            # Para birimini formatlama
            for row in ws_cari.iter_rows(min_row=2, min_col=3, max_col=5):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Kasa sayfası
            ws_kasa = wb.create_sheet("Kasa")
            ws_kasa.append(["ID", "Kasa Adı", "Gelir", "Gider", "Bakiye"])
            
            for kasa in self.kasa_listele():
                bakiye = self.kasa_bakiye_hesapla(kasa[0])
                ws_kasa.append([kasa[0], kasa[1], kasa[2], kasa[3], bakiye])
            
            # Para birimini formatlama
            for row in ws_kasa.iter_rows(min_row=2, min_col=3, max_col=5):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Faturalar sayfası
            ws_fatura = wb.create_sheet("Faturalar")
            ws_fatura.append(["ID", "Fatura No", "Tutar", "Tür"])
            
            for fatura in self.fatura_listele():
                ws_fatura.append(list(fatura))
            
            # Para birimini formatlama
            for row in ws_fatura.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Çek/Senet sayfası
            ws_cek = wb.create_sheet("Çek ve Senetler")
            ws_cek.append(["ID", "Evrak Türü", "Vade Tarihi", "Tutar"])
            
            for evrak in self.vadesi_yaklasan_cek_senetler(365):  # Tüm yılı göster
                ws_cek.append(list(evrak))
            
            # Para birimini ve tarihi formatlama
            for row in ws_cek.iter_rows(min_row=2):
                row[2].number_format = 'DD.MM.YYYY'  # Vade tarihi
                row[3].number_format = '#,##0.00 ₺'  # Tutar
            
            # Stil ayarları
            for ws in [ws_cari, ws_kasa, ws_fatura, ws_cek]:
                # Başlık satırını kalın yap ve arka plan rengini ayarla
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                
                # Sütun genişliklerini ayarla
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosya adı belirleme
            if not dosya_yolu:
                tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
                dosya_yolu = f"muhasebe_rapor_{tarih}.xlsx"
            
            # Dosyayı kaydet
            wb.save(dosya_yolu)
            return True, dosya_yolu
            
        except Exception as e:
            return False, str(e) 

    # Rapor İşlemleri
    def cari_hesap_raporu(self) -> tuple[bool, str]:
        """Cari hesapların detaylı raporunu oluşturur"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cari Hesap Raporu"
            
            # Başlıklar
            ws.append(["Müşteri Adı", "Toplam Borç", "Toplam Alacak", "Bakiye"])
            
            # Verileri çek
            self.db.cursor.execute("""
                SELECT musteri_adi, 
                       SUM(borc) as toplam_borc, 
                       SUM(alacak) as toplam_alacak,
                       SUM(alacak - borc) as bakiye
                FROM cari_hesap
                GROUP BY musteri_adi
                ORDER BY musteri_adi
            """)
            
            for row in self.db.cursor.fetchall():
                ws.append(row)
            
            # Toplam satırı
            ws.append(["TOPLAM", 
                      f"=SUM(B2:B{ws.max_row})", 
                      f"=SUM(C2:C{ws.max_row})", 
                      f"=SUM(D2:D{ws.max_row})"])
            
            # Stil ayarları
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Para birimi formatı
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Sütun genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosyayı kaydet
            tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
            dosya_adi = f"cari_hesap_raporu_{tarih}.xlsx"
            wb.save(dosya_adi)
            return True, dosya_adi
            
        except Exception as e:
            return False, str(e)
    
    def kasa_raporu(self) -> tuple[bool, str]:
        """Kasa hareketlerinin detaylı raporunu oluşturur"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Kasa Raporu"
            
            # Başlıklar
            ws.append(["Kasa Adı", "Toplam Gelir", "Toplam Gider", "Bakiye"])
            
            # Verileri çek
            self.db.cursor.execute("""
                SELECT kasa_adi, 
                       SUM(gelir) as toplam_gelir, 
                       SUM(gider) as toplam_gider,
                       SUM(gelir - gider) as bakiye
                FROM kasa_yonetimi
                GROUP BY kasa_adi
                ORDER BY kasa_adi
            """)
            
            for row in self.db.cursor.fetchall():
                ws.append(row)
            
            # Toplam satırı
            ws.append(["TOPLAM", 
                      f"=SUM(B2:B{ws.max_row})", 
                      f"=SUM(C2:C{ws.max_row})", 
                      f"=SUM(D2:D{ws.max_row})"])
            
            # Stil ve format ayarları
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Sütun genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosyayı kaydet
            tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
            dosya_adi = f"kasa_raporu_{tarih}.xlsx"
            wb.save(dosya_adi)
            return True, dosya_adi
            
        except Exception as e:
            return False, str(e)
    
    def fatura_raporu(self) -> tuple[bool, str]:
        """Faturaların detaylı raporunu oluşturur"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fatura Raporu"
            
            # Başlıklar
            ws.append(["Fatura Türü", "Fatura Adedi", "Toplam Tutar"])
            
            # Verileri çek
            self.db.cursor.execute("""
                SELECT tur, 
                       COUNT(*) as adet,
                       SUM(tutar) as toplam_tutar
                FROM faturalar_irsaliyeler
                GROUP BY tur
                ORDER BY tur
            """)
            
            for row in self.db.cursor.fetchall():
                ws.append(row)
            
            # Toplam satırı
            ws.append(["TOPLAM", 
                      f"=SUM(B2:B{ws.max_row})", 
                      f"=SUM(C2:C{ws.max_row})"])
            
            # Stil ve format ayarları
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '#,##0.00 ₺'
            
            # Sütun genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosyayı kaydet
            tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
            dosya_adi = f"fatura_raporu_{tarih}.xlsx"
            wb.save(dosya_adi)
            return True, dosya_adi
            
        except Exception as e:
            return False, str(e)
    
    def cek_senet_raporu(self) -> tuple[bool, str]:
        """Çek ve senetlerin vade analiz raporunu oluşturur"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Çek/Senet Raporu"
            
            # Başlıklar
            ws.append(["Evrak Türü", "Vade Tarihi", "Tutar", "Kalan Gün"])
            
            # Verileri çek
            self.db.cursor.execute("""
                SELECT evrak_turu,
                       vade_tarihi,
                       tutar,
                       CAST(julianday(vade_tarihi) - julianday('now') AS INTEGER) as kalan_gun
                FROM cek_senet
                ORDER BY vade_tarihi
            """)
            
            for row in self.db.cursor.fetchall():
                ws.append(row)
            
            # Toplam satırı
            ws.append(["TOPLAM", "", f"=SUM(C2:C{ws.max_row})", ""])
            
            # Stil ve format ayarları
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for row in ws.iter_rows(min_row=2):
                row[1].number_format = 'DD.MM.YYYY'  # Vade tarihi
                row[2].number_format = '#,##0.00 ₺'  # Tutar
            
            # Sütun genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Dosyayı kaydet
            tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
            dosya_adi = f"cek_senet_raporu_{tarih}.xlsx"
            wb.save(dosya_adi)
            return True, dosya_adi
            
        except Exception as e:
            return False, str(e) 